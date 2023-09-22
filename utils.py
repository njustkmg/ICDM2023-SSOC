import numpy as np
from scipy.optimize import linear_sum_assignment
from sklearn.metrics import precision_score, accuracy_score, roc_auc_score, normalized_mutual_info_score, adjusted_rand_score
import openpyxl as op
from openpyxl import Workbook
import torch
import torch.nn.functional as F
from sklearn import metrics
from scipy.optimize import linear_sum_assignment
import os

def accuracy(output, target):
    num_correct = np.sum(output == target)
    res = num_correct / len(target)
    return res


def cluster_acc(output, target):
    target = target.astype(np.int64)
    assert output.size == target.size
    D = max(output.max(), target.max()) + 1
    w = np.zeros((D, D), dtype=np.int64)
    for i in range(output.size):
        w[output[i], target[i]] += 1
    row_ind, col_ind = linear_sum_assignment(w.max() - w)
    return w[row_ind, col_ind].sum() / output.size


def entropy(x):
    EPS = 1e-8
    x_ =  torch.clamp(x, min = EPS)
    b =  x_ * torch.log(x_)

    if len(b.size()) == 2:
        return - b.sum(dim = 1).mean()
    elif len(b.size()) == 1:
        return - b.sum()
    else:
        raise ValueError('Input tensor is %d-Dimensional' %(len(b.size())))


def get_auc_roc(output, target):
    y_true = []
    for tar in target:
        lb = [0] * output.shape[1]
        lb[tar] = 1
        y_true.append(lb)
    y_true = np.array(y_true)
    return roc_auc_score(y_true, output)


def op_toexcel(data, filename):
    if not os.path.exists(filename):
        wb = Workbook()
        wb.create_sheet('Sheet')
        ws = wb['Sheet']
        ws.append(['dataset', 'batch_size', 'lr1', 'lr2', 'epoch', 'label_ratio', 'novel_ratio', 
                   'novel_classes', 'seen_acc', 'unseen_acc', 'unseen_nmi',
                   'all_acc', 'all_error', 'all_auroc'])
        wb.save(filename)

    wb = op.load_workbook(filename)
    ws = wb['Sheet']
    ws.append(data)
    wb.save(filename)


def bce_loss(output, target):
    pred = torch.sigmoid(output)
    label = torch.sigmoid(target)
    loss = -(label * torch.log(pred) + (1 - label) * torch.log(1 - pred))
    return loss

